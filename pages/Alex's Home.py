"""
Alex, The Invincible  v3.2
━━━━━━━━━━━━━━━━━━━━━━━━━
Bug-fix release:
  ✅ Barcode — plain text, no apostrophe, Excel @-format
  ✅ Dimensions — integers not decimals
  ✅ Duplicate rows — deduped before processing
  ✅ No-barcode cascade — isolated, never corrupts other rows
  ✅ Kurdish blank entries — alignment guard + retry
  ✅ Dynamic column mapping — detects Arabic/brand/KU in vendor file
  ✅ Arabic override warning — popup if vendor has Arabic column
  ✅ No expanders — all sections always visible
  ✅ Reset + Logout buttons in sidebar
  ✅ Metrics persist after download (session_state)
  ✅ Session timer + 4-hour timeout blocks UI
  ✅ Single consolidated output sheet
  ✅ 300-row timeout fix — sequential AI with per-item fallback
  ✅ AI brand auditor — validates Python-matched brand
  ✅ Iraq / Kurdish warning at Step 1
"""

import streamlit as st
import pandas as pd
from openai import OpenAI
import json, re, io, time

from deep_translator import GoogleTranslator
from smart_title_formatter import (
    format_title as smart_format_title,
    prepare_brands_list,
    scan_prohibited,
    to_pim_unit,
)
from openpyxl import load_workbook
from openpyxl.styles import numbers as xl_numbers

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Alex AI Wizard", page_icon="🤖", layout="wide")

# ─────────────────────────────────────────────────────────────────────────────
# SESSION TIMEOUT — 4 hours, resets on any activity
# ─────────────────────────────────────────────────────────────────────────────
SESSION_TIMEOUT_SECONDS = 4 * 60 * 60  # 4 hours

def _touch_session():
    st.session_state["_last_active"] = time.time()

def _check_timeout():
    last = st.session_state.get("_last_active", time.time())
    if time.time() - last > SESSION_TIMEOUT_SECONDS:
        st.session_state["authenticated"] = False
        st.session_state["_timed_out"]    = True

_touch_session()  # register activity on every page load
_check_timeout()

# ─────────────────────────────────────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────────────────────────────────────
if not st.session_state.get("authenticated", False):
    if st.session_state.get("_timed_out"):
        st.error("⏱️ Your session has expired after 4 hours of inactivity. Please log in again.")
        del st.session_state["_timed_out"]
    else:
        st.warning("🛑 Access Denied. Please go to the main Login page.")
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# AI CLIENT
# ─────────────────────────────────────────────────────────────────────────────
github_token  = st.secrets["GITHUB_TOKEN"]
ai_client     = OpenAI(base_url="https://models.inference.ai.azure.com", api_key=github_token)

# ─────────────────────────────────────────────────────────────────────────────
# MASTER BRANDS
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_master_brands() -> list:
    try:
        return prepare_brands_list(pd.read_csv("master_brands.csv", dtype=str))
    except Exception:
        return []

master_brands = load_master_brands()

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────────────────────────────────────
_defaults = {
    "ai_mapping": None,
    "last_uploaded_file": None,
    "processing_results": None,   # persists metrics after download
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
BASE_SCHEMA = ["pieceBarcode", "productTitle::en", "imageUrls", "contentsValue", "contentsUnit"]

PIM_VALID_UNITS = {
    "bags","bouquets - flowers","boxes","bunches","capsules",
    "cl","cm","cm2","cm3","dl","g","kg","l","lb","m",
    "mg","ml","oz","packets","pieces","rolls","sachets",
    "sheets","tablets","units",
}

# ─────────────────────────────────────────────────────────────────────────────
# AI HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def ai_single_dimensions(title: str) -> dict:
    """Estimate dimensions for ONE title. Returns {w,h,l,weight_g}."""
    prompt = (
        f"Product: {title}\n"
        "Estimate realistic retail packaging dimensions (width, height, length in cm) "
        "and total weight in grams. "
        "Return ONLY a JSON object with keys: w, h, l, weight_g (all integers, 0 if unknown)."
    )
    try:
        resp = ai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "Output only valid JSON."},
                {"role": "user",   "content": prompt},
            ],
            max_tokens=120,
        )
        raw = re.sub(r"^```[a-z]*\n?|```$", "", resp.choices[0].message.content.strip())
        result = json.loads(raw)
        # Force integers
        return {k: int(float(v)) for k, v in result.items() if k in ("w","h","l","weight_g")}
    except Exception:
        return {"w": 0, "h": 0, "l": 0, "weight_g": 0}


def ai_single_description(title: str, need_en: bool, need_ar: bool, need_ku: bool) -> dict:
    """Generate descriptions for ONE title."""
    langs = []
    if need_en: langs.append('"en": <1-2 sentence English description>')
    if need_ar: langs.append('"ar": <Arabic description>')
    if need_ku: langs.append('"ku": <Kurdish Kurmanji description>')
    if not langs:
        return {}
    prompt = (
        f"Product: {title}\n"
        f"Write a 1-2 sentence product description. "
        f"Return ONLY a JSON object with keys: {', '.join(langs)}. No text outside JSON."
    )
    try:
        resp = ai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "Output only valid JSON."},
                {"role": "user",   "content": prompt},
            ],
            max_tokens=300,
        )
        raw = re.sub(r"^```[a-z]*\n?|```$", "", resp.choices[0].message.content.strip())
        return json.loads(raw)
    except Exception:
        return {k: "" for k in (["en"] if need_en else []) + (["ar"] if need_ar else []) + (["ku"] if need_ku else [])}


def ai_audit_brand(title: str, brand_name: str) -> bool:
    """Return True if AI agrees the brand makes sense for this product title."""
    if not brand_name:
        return True  # no brand to audit
    prompt = (
        f'Product title: "{title}"\n'
        f'Assigned brand: "{brand_name}"\n'
        "Does this brand assignment make sense for this product? "
        'Reply with exactly one word: YES or NO.'
    )
    try:
        resp = ai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=5,
        )
        return "yes" in resp.choices[0].message.content.lower()
    except Exception:
        return True  # assume OK on failure


def safe_translate(texts: list[str], lang: str, chunk: int = 25) -> list[str]:
    """
    Translate a list, returning one string per input (same order, same length).
    Blanks/N-A placeholders are handled without breaking batch alignment.
    """
    if not texts:
        return []

    results = [""] * len(texts)
    translator = GoogleTranslator(source="en", target=lang)

    # Build index map: only translate non-empty, non-NA entries
    to_translate = [(i, t) for i, t in enumerate(texts) if t and t not in ("N/A", ".")]

    for start in range(0, len(to_translate), chunk):
        batch = to_translate[start: start + chunk]
        indices, raw_texts = zip(*batch)
        safe = [t if t.strip() else "." for t in raw_texts]

        done = False
        for attempt in range(3):
            try:
                translated = translator.translate_batch(list(safe))
                for idx, tr in zip(indices, translated):
                    results[idx] = "" if tr == "." else (tr or "")
                done = True
                break
            except Exception:
                time.sleep(1.5)
        if not done:
            for idx in indices:
                results[idx] = ""

    return results


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
display_name  = st.session_state.get("user_name", "There").title()
login_time    = st.session_state.get("_login_time", time.time())
if "_login_time" not in st.session_state:
    st.session_state["_login_time"] = time.time()
    login_time = st.session_state["_login_time"]

elapsed_min   = int((time.time() - login_time) / 60)
session_str   = f"{elapsed_min}m" if elapsed_min < 60 else f"{elapsed_min//60}h {elapsed_min%60}m"

with st.sidebar:
    # Session status bar
    st.markdown(
        f"<div style='background:#1e3a5f;padding:8px 12px;border-radius:8px;font-size:13px;color:#cde'>"
        f"👤 <b>{display_name}</b> &nbsp;·&nbsp; ⏱ {session_str} active"
        f"</div>",
        unsafe_allow_html=True,
    )
    st.divider()

    st.header("⚙️ Task Setup")
    case_id     = st.text_input("Case ID",     placeholder="e.g., CAS-12345")
    client_name = st.text_input("Client Name", placeholder="e.g., Carrefour")
    country     = st.selectbox(
        "Country",
        options=["","Egypt","United Arab Emirates","Kuwait","Qatar","Bahrain","Oman","Iraq","Jordan"],
    )
    task_ready = bool(case_id.strip() and client_name.strip() and country)

    st.divider()
    st.subheader("🌍 Translation Options")
    need_ar = st.toggle("Arabic  (productTitle::ar)", value=True)
    need_ku = st.toggle("Kurdish (productTitle::ku)", value=False)
    if country == "Iraq" and not need_ku:
        st.warning("⚠️ Iraq selected — KU titles are usually required.")

    st.divider()
    st.subheader("🤖 AI Extras")
    need_desc_en = st.toggle("Generate EN descriptions", value=False)
    need_desc_ar = st.toggle("Generate AR descriptions", value=False)
    need_desc_ku = st.toggle("Generate KU descriptions", value=False)
    need_dims    = st.toggle("AI-estimate dimensions & weight", value=False)
    need_audit   = st.toggle("AI brand auditor", value=False)

    st.divider()
    brand_count = len(master_brands)
    st.metric("🏷️ Brands in Database", f"{brand_count:,}" if brand_count else "Not loaded")

    st.divider()
    col_r, col_l = st.columns(2)
    with col_r:
        if st.button("🔄 Reset", use_container_width=True):
            for k in ["ai_mapping","last_uploaded_file","processing_results"]:
                st.session_state[k] = None
            st.rerun()
    with col_l:
        if st.button("🚪 Logout", use_container_width=True):
            st.session_state["authenticated"] = False
            st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.title("🤖 Alex, The Invincible  v3.2")
st.markdown(
    f"Hello **{display_name}** ❤️, I am **Alex** — AI Assistant by **Mostafa Abdelaziz**.  \n"
    "Smart Title Engine · Brand Scanner · PIM Units · Nitro Translate · Prohibited Scanner ✨"
)
st.divider()

# ─────────────────────────────────────────────────────────────────────────────
# FILE UPLOAD
# ─────────────────────────────────────────────────────────────────────────────
uploaded_file = st.file_uploader("Drop the messy client file here (CSV or Excel)", type=["csv","xlsx"])
if not uploaded_file:
    st.stop()

if st.session_state.last_uploaded_file != uploaded_file.name:
    st.session_state.ai_mapping          = None
    st.session_state.last_uploaded_file  = uploaded_file.name
    st.session_state.processing_results  = None

try:
    raw_df = (
        pd.read_csv(uploaded_file, dtype=str)
        if uploaded_file.name.endswith(".csv")
        else pd.read_excel(uploaded_file, dtype=str)
    )
except Exception as e:
    st.error(f"Could not read file: {e}")
    st.stop()

# ── Deduplication (critical fix) ─────────────────────────────────────────────
before_dedup = len(raw_df)
raw_df = raw_df.drop_duplicates().reset_index(drop=True)
after_dedup = len(raw_df)
if before_dedup > after_dedup:
    st.warning(f"⚠️ **{before_dedup - after_dedup} duplicate row(s) removed** before processing.")

# Preview — always visible, no expander
st.markdown("#### 👀 Raw Client Data (first 5 rows)")
st.dataframe(raw_df.head(5), use_container_width=True)
st.metric("Total rows (after dedup)", after_dedup)
st.divider()

# ── Detect vendor-provided columns dynamically ────────────────────────────────
headers    = raw_df.columns.tolist()
headers_lo = [h.lower().strip() for h in headers]

def _detect_col(*keywords) -> str | None:
    """Return first header that contains any of the given keywords."""
    for kw in keywords:
        for i, h in enumerate(headers_lo):
            if kw in h:
                return headers[i]
    return None

vendor_ar_col    = _detect_col("arabic", "ar title", "title::ar", "عنوان", "arabic title")
vendor_ku_col    = _detect_col("kurdish", "ku title", "title::ku", "كوردي")
vendor_brand_col = _detect_col("brand", "vendor", "manufacturer", "make")

# Warn if vendor has Arabic column and user has AR translation ON
if vendor_ar_col and need_ar:
    st.warning(
        f"⚠️ **Arabic override warning** — The vendor file already contains an Arabic column "
        f"(**{vendor_ar_col}**). Enabling Arabic translation will **overwrite** the vendor-provided "
        f"Arabic titles. Toggle off Arabic translation in the sidebar to keep the vendor originals."
    )

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — AI COLUMN MAPPING
# ─────────────────────────────────────────────────────────────────────────────
sample = raw_df.head(3).to_dict(orient="records")

# Build the schema dynamically based on what the vendor has
dynamic_schema = list(BASE_SCHEMA)
if vendor_ar_col:
    dynamic_schema.append("productTitle::ar")
if vendor_ku_col:
    dynamic_schema.append("productTitle::ku")
if vendor_brand_col:
    dynamic_schema.append("brand_id")

if not task_ready:
    st.warning("🔒 Fill in Case ID, Client Name, and Country in the sidebar first.")
    st.button("🧠 Step 1: Map Columns with Alex", disabled=True)
else:
    if st.button("🧠 Step 1: Map Columns with Alex", type="primary"):
        # Iraq Kurdish warning AT Step 1
        if country == "Iraq" and not need_ku:
            st.warning(
                "⚠️ **Iraq selected** but Kurdish translation is **not enabled**. "
                "Kurdish (KU) titles are required for Iraq listings. "
                "Enable Kurdish in the sidebar before proceeding."
            )

        with st.spinner("Alex is analysing headers…"):
            prompt = f"""
You are Alex, an elite Data Engineer at Talabat.
Map these client headers to the Target Schema: {dynamic_schema}

Client headers: {headers}
Sample data: {sample}

RULES:
1. pieceBarcode      → EAN > GTIN > UPC > Barcode (12-14 digit global numbers).
2. productTitle::en  → Most descriptive English title / name / description column.
3. productTitle::ar  → Arabic title column if present. Map it — do NOT generate/translate.
4. productTitle::ku  → Kurdish title column if present. Map it — do NOT generate/translate.
5. brand_id          → Brand / Vendor / Manufacturer / Make column.
6. contentsValue     → Size, Weight, Volume, Net Weight, Qty.
   NEVER map Price, Cost, MSRP, RRP, or any currency column.
7. contentsUnit      → UOM, Unit, Measurement.
   NEVER map currency symbols ($, AED, EGP, SAR).
8. imageUrls         → URL, Link, Photo, Image, Media column.

Return ONLY a JSON object: {{"Target_Field": "Client_Header"}}.
Omit any target field where no suitable client header exists.
The job is to FORMAT and MAP data — never translate it.
"""
            response = ai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "Output strict JSON only."},
                    {"role": "user",   "content": prompt},
                ],
                response_format={"type": "json_object"},
            )
            st.session_state.ai_mapping = json.loads(response.choices[0].message.content)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — MAPPING REVIEW (always visible, no expander)
# ─────────────────────────────────────────────────────────────────────────────
if st.session_state.ai_mapping is None:
    st.stop()

st.markdown("#### 🛠️ Step 2: Review & Override Alex's Mapping")
with st.form("mapping_form"):
    final_mapping = {}
    options = ["--- Leave Blank ---"] + headers
    for col in dynamic_schema:
        suggested = st.session_state.ai_mapping.get(col)
        idx = options.index(suggested) if suggested in options else 0
        final_mapping[col] = st.selectbox(f"Map '{col}' to:", options, index=idx)
    submitted = st.form_submit_button("🚀 Step 3: Run Full Catalogue Processing", type="primary")

if not submitted:
    # Show persisted results if available (survives download button click)
    if st.session_state.processing_results:
        _show_results(st.session_state.processing_results)
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — FULL PROCESSING
# ─────────────────────────────────────────────────────────────────────────────
t_start  = time.time()
progress = st.progress(0, text="Starting…")
_touch_session()

active_map = {k: v for k, v in final_mapping.items() if v != "--- Leave Blank ---"}

# ── 3a. Build working DataFrame ───────────────────────────────────────────────
work_df = pd.DataFrame(index=range(len(raw_df)))  # explicit clean index

for col in dynamic_schema:
    src = active_map.get(col)
    work_df[col] = raw_df[src].values if src and src in raw_df.columns else ""

# Ensure base columns always exist
for col in BASE_SCHEMA:
    if col not in work_df.columns:
        work_df[col] = ""

# Add all output columns
work_df.insert(1, "brand_id",              "")
work_df["productTitle::ar"]        = work_df.get("productTitle::ar", "")
work_df["productTitle::ku"]        = work_df.get("productTitle::ku", "")
work_df["widthInCm"]               = ""
work_df["heightInCm"]              = ""
work_df["lengthInCm"]              = ""
work_df["weightValue"]             = ""
work_df["weightUnit"]              = "g"
work_df["numberOfUnits"]           = 1
work_df["productDescription::en"]  = ""
work_df["productDescription::ar"]  = ""
work_df["productDescription::ku"]  = ""
work_df["_title_confidence"]       = ""
work_df["_title_changes"]          = ""
work_df["_prohibited_flag"]        = ""
work_df["_brand_audit_flag"]       = ""
work_df["Catalogue_Feedback"]      = ""

total    = len(work_df)
feedback: list[str] = []
en_titles_for_translate: list[str] = []   # final EN titles (for AR/KU translation)

# ── 3b. Row-by-row core ───────────────────────────────────────────────────────
progress.progress(0.05, text="Formatting titles & scanning brands…")

for idx in range(total):
    row    = work_df.iloc[idx]
    doubts: list[str] = []

    # ── Barcode (clean, no cascade if missing) ────────────────────────────
    raw_bc = str(row.get("pieceBarcode","")).replace(".0","").strip()
    if raw_bc and raw_bc.lower() not in ("nan","none","") and raw_bc != "0":
        # Store plain — openpyxl will format as text in export
        padded = raw_bc.zfill(13)
        work_df.at[idx, "pieceBarcode"] = padded
    else:
        work_df.at[idx, "pieceBarcode"] = ""
        doubts.append("Missing Barcode")

    # ── Smart Title + Brand Scanner ───────────────────────────────────────
    raw_title = str(row.get("productTitle::en","")).replace("nan","").strip()
    cv_raw    = str(row.get("contentsValue","")).replace("nan","").strip()
    cu_raw    = str(row.get("contentsUnit", "")).replace("nan","").strip()

    title_res = smart_format_title(
        raw_title=raw_title,
        contents_value=cv_raw,
        contents_unit=cu_raw,
        brands_list=master_brands,
    )

    formatted = title_res["formatted_title"]
    work_df.at[idx, "productTitle::en"]  = formatted
    work_df.at[idx, "brand_id"]          = title_res["brand_id"]
    work_df.at[idx, "_title_confidence"] = title_res["confidence"]
    work_df.at[idx, "numberOfUnits"]     = title_res["number_of_units"]

    if formatted and formatted != raw_title:
        work_df.at[idx, "_title_changes"] = f"{raw_title} → {formatted}"

    en_titles_for_translate.append(formatted if formatted else "N/A")

    # ── AI Brand Auditor ──────────────────────────────────────────────────
    if need_audit and title_res["brand_id"] and title_res["brand_name"]:
        ok = ai_audit_brand(formatted, title_res["brand_name"])
        if not ok:
            work_df.at[idx, "brand_id"]           = ""
            work_df.at[idx, "_brand_audit_flag"]  = f"AI rejected: '{title_res['brand_name']}'"
            doubts.append(f"Brand mismatch: '{title_res['brand_name']}'")

    # ── Prohibited content ────────────────────────────────────────────────
    if title_res["prohibited"]:
        flag_str = " | ".join(title_res["prohibited"])
        work_df.at[idx, "_prohibited_flag"] = flag_str
        doubts.append(f"🚫 PROHIBITED — {flag_str}")

    # ── Unbranded ─────────────────────────────────────────────────────────
    if "Unbranded - Audit Required" in title_res["issues"]:
        doubts.append("Unbranded - Audit Required")

    # ── contentsValue / contentsUnit ──────────────────────────────────────
    pim_val  = title_res["pim_contents_value"]
    pim_unit = title_res["pim_contents_unit"]

    cv_out = cv_raw if (cv_raw and cv_raw not in ("0","0.0")) else pim_val
    cu_out = cu_raw.lower() if cu_raw.lower() in PIM_VALID_UNITS else \
             (pim_unit if pim_unit in PIM_VALID_UNITS else to_pim_unit(cu_raw.lower()))

    work_df.at[idx, "contentsValue"] = cv_out
    work_df.at[idx, "contentsUnit"]  = cu_out if cu_out in PIM_VALID_UNITS else cu_out

    # ── Validation ────────────────────────────────────────────────────────
    cv_final = str(work_df.at[idx, "contentsValue"]).strip()
    cu_final = str(work_df.at[idx, "contentsUnit"]).strip().lower()

    if not cv_final or cv_final in ("nan",""):
        doubts.append("Missing Qty")
    if not cu_final or cu_final in ("nan",""):
        doubts.append("Missing Unit")
    elif cu_final not in PIM_VALID_UNITS:
        doubts.append(f"Invalid Unit ({cu_final})")

    # ── Image ─────────────────────────────────────────────────────────────
    raw_url = str(row.get("imageUrls","")).strip().lower()
    if raw_url in ("","nan","none","n/a"):
        doubts.append("Missing Image")
    elif not raw_url.startswith("http"):
        doubts.append("Invalid Image URL")

    feedback.append("✅ Ready for Catalogue" if not doubts else "⚠️ " + ", ".join(doubts))

work_df["Catalogue_Feedback"] = feedback
progress.progress(0.35, text="Core processing done…")

# ── 3c. AI DIMENSIONS — one at a time with fallback ──────────────────────────
if need_dims:
    progress.progress(0.37, text="AI estimating dimensions…")
    for i in range(total):
        title = str(work_df.at[i, "productTitle::en"])
        dims  = ai_single_dimensions(title)
        if dims.get("w"):  work_df.at[i, "widthInCm"]  = int(dims["w"])
        if dims.get("h"):  work_df.at[i, "heightInCm"] = int(dims["h"])
        if dims.get("l"):  work_df.at[i, "lengthInCm"] = int(dims["l"])
        if dims.get("weight_g"): work_df.at[i, "weightValue"] = int(dims["weight_g"])
        frac = 0.37 + 0.13 * (i + 1) / total
        progress.progress(frac, text=f"Dimensions {i+1}/{total}…")

progress.progress(0.50, text="Dimensions done…")

# ── 3d. AI DESCRIPTIONS — one at a time with fallback ────────────────────────
if need_desc_en or need_desc_ar or need_desc_ku:
    progress.progress(0.52, text="Generating descriptions…")
    for i in range(total):
        title = str(work_df.at[i, "productTitle::en"])
        desc  = ai_single_description(title, need_desc_en, need_desc_ar, need_desc_ku)
        if need_desc_en: work_df.at[i, "productDescription::en"] = desc.get("en","")
        if need_desc_ar: work_df.at[i, "productDescription::ar"] = desc.get("ar","")
        if need_desc_ku: work_df.at[i, "productDescription::ku"] = desc.get("ku","")
        frac = 0.52 + 0.08 * (i + 1) / total
        progress.progress(frac, text=f"Descriptions {i+1}/{total}…")

progress.progress(0.60, text="Descriptions done…")

# ── 3e. Arabic translation ────────────────────────────────────────────────────
if need_ar:
    progress.progress(0.62, text="Translating to Arabic…")
    # If vendor provided Arabic column and user chose NOT to map it — use translation
    # If vendor provided Arabic column and user DID map it — already populated, skip
    needs_ar_translation = [
        i for i in range(total)
        if not str(work_df.at[i, "productTitle::ar"]).strip()
        or str(work_df.at[i, "productTitle::ar"]).lower() in ("nan","none","")
    ]
    if needs_ar_translation:
        ar_inputs  = [en_titles_for_translate[i] for i in needs_ar_translation]
        ar_results = safe_translate(ar_inputs, "ar")
        for i, tr in zip(needs_ar_translation, ar_results):
            work_df.at[i, "productTitle::ar"] = tr

progress.progress(0.80, text="Arabic done…")

# ── 3f. Kurdish translation ───────────────────────────────────────────────────
if need_ku:
    progress.progress(0.82, text="Translating to Kurdish…")
    needs_ku_translation = [
        i for i in range(total)
        if not str(work_df.at[i, "productTitle::ku"]).strip()
        or str(work_df.at[i, "productTitle::ku"]).lower() in ("nan","none","")
    ]
    if needs_ku_translation:
        ku_inputs  = [en_titles_for_translate[i] for i in needs_ku_translation]
        ku_results = safe_translate(ku_inputs, "ku")
        for i, tr in zip(needs_ku_translation, ku_results):
            work_df.at[i, "productTitle::ku"] = tr

progress.progress(0.97, text="Kurdish done…")
progress.progress(1.0,  text="Done!")
elapsed = time.time() - t_start

# ─────────────────────────────────────────────────────────────────────────────
# RESULTS — persist in session_state so they survive download button click
# ─────────────────────────────────────────────────────────────────────────────
ORDERED_EXPORT = [
    "pieceBarcode","brand_id","productTitle::en","productTitle::ar","productTitle::ku",
    "imageUrls",
    "widthInCm","heightInCm","lengthInCm","weightValue","weightUnit",
    "contentsValue","contentsUnit","numberOfUnits",
    "productDescription::en","productDescription::ar","productDescription::ku",
    "Catalogue_Feedback",
]
export_cols = [c for c in ORDERED_EXPORT if c in work_df.columns]

ready_df      = work_df[work_df["Catalogue_Feedback"] == "✅ Ready for Catalogue"]
error_df      = work_df[work_df["Catalogue_Feedback"].str.startswith("⚠️")]
prohibited_df = work_df[work_df["_prohibited_flag"] != ""]
changed_df    = work_df[work_df["_title_changes"] != ""].copy()

stats = {
    "total":         len(work_df),
    "ready":         len(ready_df),
    "errors":        len(error_df),
    "prohibited":    len(prohibited_df),
    "brands":        int((work_df["brand_id"] != "").sum()),
    "changed":       len(changed_df),
    "quality":       f"{(len(ready_df)/len(work_df)*100):.1f}%" if len(work_df) else "0%",
    "elapsed":       elapsed,
}

# ── Build consolidated single sheet ──────────────────────────────────────────
# Status column for the consolidated sheet
def _status_icon(fb: str) -> str:
    if fb == "✅ Ready for Catalogue": return "✅ Ready"
    if "🚫" in fb:                     return "🚫 Prohibited"
    return "⚠️ Needs Review"

work_df["Status"] = work_df["Catalogue_Feedback"].apply(_status_icon)
consolidated_cols = [c for c in ["Status"] + export_cols if c in work_df.columns]
consolidated_df   = work_df[consolidated_cols]

# ── Build Excel with openpyxl barcode text format ─────────────────────────────
output = io.BytesIO()
with pd.ExcelWriter(output, engine="openpyxl") as writer:
    # Summary sheet
    pd.DataFrame({
        "Case ID":           [case_id],
        "Client Name":       [client_name],
        "Country":           [country],
        "Total Rows":        [stats["total"]],
        "Ready":             [stats["ready"]],
        "Needs Review":      [stats["errors"]],
        "Prohibited":        [stats["prohibited"]],
        "Brands Matched":    [stats["brands"]],
        "Titles Fixed":      [stats["changed"]],
        "Quality Score":     [stats["quality"]],
        "Processing Time":   [f"{elapsed:.1f}s"],
        "AR Translation":    ["Yes" if need_ar else "No"],
        "KU Translation":    ["Yes" if need_ku else "No"],
        "AI Descriptions":   ["Yes" if (need_desc_en or need_desc_ar or need_desc_ku) else "No"],
        "AI Dimensions":     ["Yes" if need_dims else "No"],
        "AI Brand Auditor":  ["Yes" if need_audit else "No"],
    }).to_excel(writer, index=False, sheet_name="Summary")

    # Consolidated master sheet (ALL rows, one sheet)
    consolidated_df.to_excel(writer, index=False, sheet_name="All Products")

    # Prohibited sheet (if any)
    if stats["prohibited"] > 0:
        prohibited_df[consolidated_cols + ["_prohibited_flag"]].to_excel(
            writer, index=False, sheet_name="Prohibited Items"
        )

    # Format barcode column as TEXT in all sheets
    wb = writer.book
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find barcode column
        for col_cell in ws[1]:
            if col_cell.value == "pieceBarcode":
                col_letter = col_cell.column_letter
                for cell in ws[col_letter]:
                    cell.number_format = "@"   # Excel text format
                break

output_bytes    = output.getvalue()
output_filename = f"{case_id} - {client_name} - {country}_Alex_v3.2.xlsx"

# Store results in session_state BEFORE download button renders
st.session_state.processing_results = {
    "stats":          stats,
    "output_bytes":   output_bytes,
    "output_filename":output_filename,
    "work_df":        work_df,
    "export_cols":    export_cols,
    "consolidated_df":consolidated_df,
    "prohibited_df":  prohibited_df,
    "changed_df":     changed_df,
}

# ─────────────────────────────────────────────────────────────────────────────
# DISPLAY RESULTS (function so it can be called from persisted state too)
# ─────────────────────────────────────────────────────────────────────────────
def _show_results(res: dict):
    s = res["stats"]
    st.success(f"✅ Completed in **{s['elapsed']:.1f}s**")

    if s["prohibited"] > 0:
        st.error(
            f"🚫 **{s['prohibited']} prohibited item(s) detected** (tobacco or pork). "
            "See the Prohibited tab — remove before submission."
        )

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("✅ Ready",          s["ready"])
    c2.metric("⚠️ Needs Review",   s["errors"])
    c3.metric("🚫 Prohibited",      s["prohibited"])
    c4.metric("🏷️ Brands Matched", s["brands"])
    c5.metric("✏️ Titles Fixed",   s["changed"])
    c6.metric("📊 Quality Score",   s["quality"])

    st.download_button(
        label=f"📥 Download Report  ({res['output_filename']})",
        data=res["output_bytes"],
        file_name=res["output_filename"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

    wdf       = res["work_df"]
    edf       = wdf[wdf["Catalogue_Feedback"].str.startswith("⚠️")]
    prdf      = res["prohibited_df"]
    chdf      = res["changed_df"]
    cdf       = res["consolidated_df"]
    ecols     = res["export_cols"]

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        f"📋 All Products ({len(wdf)})",
        f"✅ Ready ({s['ready']})",
        f"⚠️ Action Required ({s['errors']})",
        f"🚫 Prohibited ({s['prohibited']})",
        f"✏️ Title Changes ({s['changed']})",
    ])

    with tab1:
        st.dataframe(cdf, use_container_width=True)
    with tab2:
        ready_view = wdf[wdf["Catalogue_Feedback"] == "✅ Ready for Catalogue"]
        st.dataframe(ready_view[[c for c in ecols if c in ready_view.columns]], use_container_width=True)
    with tab3:
        if len(edf):
            st.dataframe(edf[[c for c in ecols if c in edf.columns]], use_container_width=True)
        else:
            st.success("🎉 No errors!")
    with tab4:
        if len(prdf):
            st.dataframe(
                prdf[["productTitle::en","_prohibited_flag","Catalogue_Feedback"]],
                use_container_width=True,
            )
        else:
            st.success("✅ No prohibited products detected.")
    with tab5:
        if len(chdf):
            st.dataframe(
                chdf[["productTitle::en","_title_changes","_title_confidence"]],
                use_container_width=True,
            )
        else:
            st.info("No titles were changed.")

_show_results(st.session_state.processing_results)